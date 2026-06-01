"""
excel_parser.py — Parses the CSO travel Excel file and returns
clean payroll records ready to append to the master sheet.

Supports TWO sheet layouts:

LAYOUT A — "Present absent" (original single-block format):
  Row 1-4 : Date headers (Row 4 contains datetime objects or 'DD_Mon_YY' strings)
  Row 5   : Column headers ('Employee ID', 'Employee', …)
  Row 6+  : One employee per row

LAYOUT B — "Extra km" (multi-block stacked format):
  Multiple payroll periods are stacked vertically in one sheet.
  Each block looks like:
    Row N  : Day names  (Saturday, Sunday, …)
    Row N+1: Dates      (datetime objects OR 'DD_Mon_YY' strings, cols 5+)
    Row N+2: Headers    ('Employee ID', 'Employee', 'Designation', …)
    Row N+3: (blank — optional)
    Row N+4+: Employee data rows
    (blank rows)
    … next block …

The parser auto-detects the layout, extracts all blocks, and returns
the same list[dict] that the downstream master_sheet.py expects.

Period detection:
  - Primary  : actual datetime objects in the dates row (openpyxl native)
  - Secondary: string dates formatted as 'DD_Mon_YY' / 'D_Mon_YY' (e.g. '20_Dec_25')
  - Fallback : filename / email-subject text scan
"""

import io
import re
import logging
from datetime import datetime, date
import openpyxl

log = logging.getLogger(__name__)

# ── Column indices (0-based) ───────────────────────────────────────────────────
COL_EMP_ID    = 0   # A
COL_EMP_NAME  = 1   # B
COL_DESIG     = 2   # C
COL_ASE       = 3   # D
COL_ASM       = 4   # E
COL_KM_START  = 5   # F  — first daily KM column
COL_KM_END    = 34  # AI — last daily KM column
COL_EXTRA_KM  = 35  # AJ — "Exception Kms in the Month"
COL_APPROVAL  = 36  # AK — "SM Approval"

RATE_PER_KM   = 3
DATA_START_ROW = 6   # 1-based (used only for legacy single-block layout)

# Abbreviated month name → month number
_MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3,  "apr": 4,
    "may": 5, "jun": 6, "jul": 7,  "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MONTH_FULL = {
    1: "January", 2: "February",  3: "March",    4: "April",
    5: "May",     6: "June",      7: "July",      8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


class ExcelParser:

    # ═══════════════════════════════════════════════════════════════════════════
    # Public entry point
    # ═══════════════════════════════════════════════════════════════════════════

    def parse(
        self,
        file_bytes: bytes,
        filename: str,
        email_subject: str,
        sender_email: str,
    ) -> list[dict]:
        """
        Parse an Excel file from raw bytes.

        Multi-sheet support: iterates over every sheet in the workbook.
        Sheets that look like payroll data (contain an 'Employee ID' header)
        are processed; purely summary/metadata sheets are skipped.
        Records from all sheets are combined and returned together.

        Within each sheet, handles both single-block and multi-block
        (stacked) layouts automatically.
        """
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception as e:
            raise ValueError(f"Cannot open workbook: {e}")

        sheets_to_parse = self._select_sheets(wb, filename)
        if not sheets_to_parse:
            raise ValueError(f"Could not find any usable sheets in {filename}")

        all_records: list[dict] = []
        for sheet_name, sheet in sheets_to_parse:
            log.info(f"  Processing sheet: '{sheet_name}'")
            sheet_records = self._parse_sheet(
                sheet, sheet_name, filename, email_subject, sender_email
            )
            all_records.extend(sheet_records)

        log.info(
            f"  Grand total: {len(all_records)} record(s) across "
            f"{len(sheets_to_parse)} sheet(s) in '{filename}'"
        )
        return all_records

    def _parse_sheet(
        self,
        sheet,
        sheet_name: str,
        filename: str,
        email_subject: str,
        sender_email: str,
    ) -> list[dict]:
        """
        Parse all payroll blocks within a single worksheet.
        Returns a list of payroll record dicts (may be empty if the sheet
        has no 'Employee ID' header row).
        """
        # Load all rows once — required for multi-block scanning
        all_rows = list(sheet.iter_rows(values_only=True))

        # Locate every payroll block (header row = 'Employee ID' at col 0)
        block_header_indices = [
            idx for idx, row in enumerate(all_rows)
            if row and str(row[0] or "").strip().lower() == "employee id"
        ]

        if not block_header_indices:
            log.warning(f"    No 'Employee ID' header rows found in sheet '{sheet_name}' of {filename}")
            return []

        records = []
        for block_num, hdr_idx in enumerate(block_header_indices):

            # ── Validate the 5 fixed columns ──────────────────────────────────
            hdr_row = all_rows[hdr_idx]
            if len(hdr_row) < 5:
                log.warning(f"    Sheet '{sheet_name}' Block {block_num+1}: too few columns at row {hdr_idx+1}, skipping.")
                continue
            c0, c1, c2, c3, c4 = (str(v or "").lower() for v in hdr_row[:5])
            if not (
                "employee id"  in c0
                and "employee" in c1
                and "designation" in c2
                and "ase manager" in c3
                and "asm manager" in c4
            ):
                log.warning(
                    f"    Sheet '{sheet_name}' Block {block_num+1}: column headers don't match "
                    f"expected format at row {hdr_idx+1}, skipping."
                )
                continue

            # ── Find the dates row (1–3 rows above the header row) ────────────
            dates_row: tuple = ()
            for back in range(1, 4):
                candidate_idx = hdr_idx - back
                if candidate_idx < 0:
                    break
                candidate = all_rows[candidate_idx]
                if self._is_dates_row(candidate):
                    dates_row = candidate
                    break

            if not dates_row:
                log.warning(
                    f"    Sheet '{sheet_name}' Block {block_num+1}: no dates row found above "
                    f"header row {hdr_idx+1} — period will fall back to filename/subject."
                )

            month_tag = self._extract_month_from_row4(dates_row, filename, email_subject)

            # ── Identify daily KM column indices for this block ───────────────
            daily_col_indices = self._find_daily_cols(dates_row, hdr_row)

            # ── Determine end of this block's data rows ───────────────────────
            # Data ends at the row before the NEXT block's dates/header row,
            # or at the end of the sheet.
            if block_num + 1 < len(block_header_indices):
                next_hdr_idx = block_header_indices[block_num + 1]
                # The dates row for the next block is 1-3 rows before its header,
                # so end data at the row just before that.
                data_end_idx = next_hdr_idx - 3  # conservative: leave room for day-names + dates
            else:
                data_end_idx = len(all_rows)

            # ── Parse employee rows ───────────────────────────────────────────
            block_added = 0
            block_skipped = 0
            for row_idx in range(hdr_idx + 1, data_end_idx):
                row = all_rows[row_idx]
                if all(v is None for v in row):
                    continue  # blank spacer row — skip, don't stop

                # Stop if we accidentally wander into a day-names or dates row
                if self._is_dates_row(row) or self._is_day_names_row(row):
                    break

                record = self._parse_row(
                    row, row_idx + 1,
                    filename, email_subject,
                    month_tag, daily_col_indices, sender_email,
                )
                if record:
                    records.append(record)
                    block_added += 1
                else:
                    block_skipped += 1

            log.info(
                f"    Sheet '{sheet_name}' Block {block_num+1} ({month_tag}): "
                f"{block_added} records, {block_skipped} skipped."
            )

        log.info(
            f"    Sheet '{sheet_name}': {len(records)} record(s) from "
            f"{len(block_header_indices)} block(s)"
        )
        return records

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet finder
    # ═══════════════════════════════════════════════════════════════════════════

    def _select_sheets(self, wb, filename: str) -> list[tuple[str, object]]:
        """
        Return an ordered list of (sheet_name, worksheet) pairs to parse.

        Strategy:
          1. Collect ALL sheets that contain at least one 'Employee ID' header
             row — these are definitively payroll sheets regardless of name.
          2. If none found via content scan, fall back to the original
             name-priority heuristic (present/absent → extra km → first
             non-empty) so behaviour is unchanged for single-sheet files.

        Sheets are returned in workbook tab order.
        """
        payroll_sheets: list[tuple[str, object]] = []

        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(max_row=200, values_only=True):
                if row and str(row[0] or "").strip().lower() == "employee id":
                    payroll_sheets.append((name, ws))
                    log.info(f"  Sheet '{name}' identified as payroll data (contains 'Employee ID' header)")
                    break  # found — no need to scan further rows in this sheet

        if payroll_sheets:
            return payroll_sheets

        # ── Fallback: no sheet had an 'Employee ID' header; use name heuristics ──
        log.warning(
            f"No sheet in '{filename}' contains an 'Employee ID' header — "
            "falling back to name-based sheet selection."
        )

        # Priority 1: exact canonical name
        if "Present absent" in wb.sheetnames:
            return [("Present absent", wb["Present absent"])]

        # Priority 2: name contains 'present' or 'absent'
        for name in wb.sheetnames:
            nl = name.lower()
            if "present" in nl or "absent" in nl:
                log.warning(f"  Using sheet '{name}' as 'Present absent' fallback in {filename}")
                return [(name, wb[name])]

        # Priority 3: name looks like an extra-km / travel sheet
        for name in wb.sheetnames:
            nl = name.lower()
            if ("extra" in nl and "km" in nl) or "travel" in nl or "km" in nl:
                log.warning(f"  Using sheet '{name}' as travel/KM sheet in {filename}")
                return [(name, wb[name])]

        # Priority 4: first non-empty sheet
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(max_row=5, values_only=True):
                if any(v is not None for v in row):
                    log.warning(f"  Using first non-empty sheet '{name}' in {filename}")
                    return [(name, ws)]

        log.warning(f"  No usable sheet found in {filename}, using wb.active")
        return [(wb.active.title, wb.active)]

    # ═══════════════════════════════════════════════════════════════════════════
    # Row type detection helpers
    # ═══════════════════════════════════════════════════════════════════════════

    def _is_dates_row(self, row: tuple) -> bool:
        """True if the row contains datetime objects or 'DD_Mon_YY' strings in cols 5+."""
        if not row or len(row) <= 5:
            return False
        hits = 0
        for v in row[5:]:
            if isinstance(v, (datetime, date)):
                hits += 1
            elif isinstance(v, str) and self._parse_string_date(v) is not None:
                hits += 1
            if hits >= 2:
                return True
        return False

    def _is_day_names_row(self, row: tuple) -> bool:
        """True if cols 5+ are mostly weekday name strings (Saturday, Sunday, …)."""
        if not row or len(row) <= 5:
            return False
        day_names = {
            "monday", "tuesday", "wednesday", "thursday",
            "friday", "saturday", "sunday",
        }
        hits = sum(
            1 for v in row[5:]
            if isinstance(v, str) and v.strip().lower() in day_names
        )
        return hits >= 3

    # ═══════════════════════════════════════════════════════════════════════════
    # Daily KM column detection (per block)
    # ═══════════════════════════════════════════════════════════════════════════

    def _find_daily_cols(self, dates_row: tuple, header_row: tuple) -> list[int]:
        """
        Return 0-based column indices that represent daily KM entries.
        Logic: col must have a date value in dates_row OR a 'km' label in header_row,
               AND must NOT be the 'Exception Kms' or 'SM Approval' summary columns.
        """
        max_cols = max(len(dates_row), len(header_row))
        daily_col_indices = []

        for col_idx in range(5, max_cols):
            val_date = dates_row[col_idx]   if col_idx < len(dates_row)  else None
            val_hdr  = header_row[col_idx]  if col_idx < len(header_row) else None

            hdr_lower = str(val_hdr or "").strip().lower()

            # Exclude summary / approval columns
            if "exception" in hdr_lower or "approval" in hdr_lower:
                break  # these are the last meaningful columns; stop scanning

            is_date = (
                isinstance(val_date, (datetime, date))
                or (isinstance(val_date, str) and self._parse_string_date(val_date) is not None)
            )
            is_km = hdr_lower in {"km's", "kms", "km", "km.s"}

            if is_date or is_km:
                daily_col_indices.append(col_idx)

        if not daily_col_indices:
            log.warning("Could not detect daily KM columns — falling back to cols 5–34")
            daily_col_indices = list(range(5, 35))

        return daily_col_indices

    # ═══════════════════════════════════════════════════════════════════════════
    # Row parser (unchanged logic, same output dict)
    # ═══════════════════════════════════════════════════════════════════════════

    def _parse_row(
        self,
        row: tuple,
        row_idx: int,
        filename: str,
        email_subject: str,
        month_tag: str,
        daily_col_indices: list[int],
        sender_email: str,
    ) -> dict | None:

        emp_id   = self._safe_str(row, COL_EMP_ID)
        emp_name = self._safe_str(row, COL_EMP_NAME)
        if not emp_id or not emp_name:
            return None

        designation = self._safe_str(row, COL_DESIG)
        ase_manager = self._safe_str(row, COL_ASE)
        asm_manager = self._safe_str(row, COL_ASM)

        # Sum dynamically identified daily KM columns
        daily_kms      = [self._to_number(row[c] if c < len(row) else None) for c in daily_col_indices]
        total_extra_km = sum(daily_kms)
        amount_inr     = total_extra_km * RATE_PER_KM

        return {
            "Employee ID":    emp_id,
            "Employee Name":  emp_name,
            "Designation":    designation,
            "ASE Manager":    ase_manager,
            "ASM Manager":    asm_manager,
            "Month":          month_tag,
            "Total Extra KM": total_extra_km,
            "Amount INR":     amount_inr,
            "Source File":    filename,
            "Email Subject":  email_subject,
            "Sender Email":   sender_email,
            "Processed Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    # ═══════════════════════════════════════════════════════════════════════════
    # Period / month-tag detection
    # ═══════════════════════════════════════════════════════════════════════════

    def _extract_month_from_row4(self, row_4: tuple, filename: str, subject: str) -> str:
        """
        Derive a human-readable period tag such as 'December - January'.

        Priority:
          1. Native datetime / date objects in row_4  (openpyxl standard)
          2. String dates in 'DD_Mon_YY' / 'D_Mon_YY' format  (NEW)
          3. Filename / email-subject text scan  (original fallback)
        """
        if not row_4:
            return self._extract_month(filename, subject)

        # ── 1. Native datetime objects ─────────────────────────────────────────
        native_dates = [v for v in row_4 if isinstance(v, (datetime, date))]
        if native_dates:
            return self._tag_from_date_list(native_dates, filename)

        # ── 2. String dates like '20_Dec_25', '1_jan_26' ──────────────────────
        parsed = []
        for v in row_4:
            if not isinstance(v, str):
                continue
            result = self._parse_string_date(v)
            if result:
                parsed.append(result)

        if parsed:
            parsed.sort()
            first_year, first_month = parsed[0]
            last_year,  last_month  = parsed[-1]
            if first_month == last_month and first_year == last_year:
                # Single month — infer previous month as start
                prev_month = 12 if first_month == 1 else first_month - 1
                tag = f"{_MONTH_FULL[prev_month]} - {_MONTH_FULL[first_month]}"
            else:
                tag = f"{_MONTH_FULL[first_month]} - {_MONTH_FULL[last_month]}"
            log.info(f"    Period detected from Row string dates: {tag} ('{filename}')")
            return tag

        # ── 3. Filename / subject fallback ────────────────────────────────────
        log.warning(
            f"No date values found in dates row of '{filename}' — "
            "falling back to filename/subject period detection."
        )
        return self._extract_month(filename, subject)

    def _tag_from_date_list(self, dates: list, filename: str) -> str:
        """Build a 'Month1 - Month2' tag from a list of date/datetime objects."""
        dates_sorted = sorted(dates, key=lambda d: (d.year, d.month) if isinstance(d, date) else (d.date().year, d.date().month))
        first = dates_sorted[0]
        last  = dates_sorted[-1]
        fm = first.month if isinstance(first, date) else first.date().month
        lm = last.month  if isinstance(last,  date) else last.date().month
        if fm == lm:
            prev = 12 if fm == 1 else fm - 1
            tag = f"{_MONTH_FULL[prev]} - {_MONTH_FULL[fm]}"
        else:
            tag = f"{_MONTH_FULL[fm]} - {_MONTH_FULL[lm]}"
        log.info(f"    Period detected from Row 4 dates: {tag} ('{filename}')")
        return tag

    def _parse_string_date(self, value: str):
        """
        Parse a string like '20_Dec_25' or '1_jan_26' into (year, month).
        Returns None if the string doesn't match the expected pattern.
        """
        parts = value.strip().replace("-", "_").split("_")
        if len(parts) < 3:
            return None
        month_str = parts[1].lower()[:3]
        year_str  = parts[2]
        if month_str not in _MONTH_ABBR:
            return None
        try:
            year = int(year_str) + 2000 if len(year_str) <= 2 else int(year_str)
            return (year, _MONTH_ABBR[month_str])
        except ValueError:
            return None

    def _extract_month(self, filename: str, subject: str) -> str:
        """
        Extract a period tag like 'March - April' from filename or email subject.
        (Original fallback — unchanged.)
        """
        text = f"{filename} {subject}".lower()

        month_order = list(_MONTH_ABBR.keys())
        matches = []
        for m in _MONTH_ABBR:
            for match in re.finditer(rf"\b{m}[a-z]*\b", text):
                matches.append((match.start(), m))

        matches.sort(key=lambda x: x[0])
        unique_months = []
        for _, m in matches:
            if m not in unique_months:
                unique_months.append(m)

        if len(unique_months) >= 2:
            m_a, m_b = unique_months[0], unique_months[1]
            idx_a = month_order.index(m_a)
            idx_b = month_order.index(m_b)
            if (idx_b - idx_a) % 12 == 1:
                m1, m2 = m_a, m_b
            elif (idx_a - idx_b) % 12 == 1:
                m1, m2 = m_b, m_a
            else:
                m1, m2 = (m_a, m_b) if idx_a < idx_b else (m_b, m_a)
            return f"{_MONTH_FULL[_MONTH_ABBR[m1]]} - {_MONTH_FULL[_MONTH_ABBR[m2]]}"

        elif len(unique_months) == 1:
            m2 = unique_months[0]
            m2_idx = month_order.index(m2)
            m1_idx = (m2_idx - 1) % 12
            m1 = month_order[m1_idx]
            return f"{_MONTH_FULL[_MONTH_ABBR[m1]]} - {_MONTH_FULL[_MONTH_ABBR[m2]]}"

        m1_idx = datetime.now().month - 1
        m2_idx = (m1_idx + 1) % 12
        fallback = (
            f"{_MONTH_FULL[_MONTH_ABBR[month_order[m1_idx]]]} - "
            f"{_MONTH_FULL[_MONTH_ABBR[month_order[m2_idx]]]}"
        )
        log.warning(f"Could not extract month from '{text}' — using {fallback}")
        return fallback

    # ═══════════════════════════════════════════════════════════════════════════
    # Low-level helpers (unchanged)
    # ═══════════════════════════════════════════════════════════════════════════

    def _safe_str(self, row: tuple, idx: int) -> str:
        if idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def _to_number(self, val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, TypeError):
            return 0.0